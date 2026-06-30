# -*- coding: utf-8 -*-
"""
Tennis technical-tactical determinant identification: a reproducible multi-model ensemble pipeline (random_state=42).
Steps: data cleaning, de-collinearization, twelve base learners, exhaustive ensemble selection by out-of-fold F1, and multi-level interpretation (native importance, feature ablation, ensemble SHAP).
"""
import warnings
from itertools import combinations
import numpy as np
import pandas as pd
from scipy.stats import shapiro, spearmanr, pearsonr
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.pipeline import Pipeline
from sklearn.model_selection import StratifiedKFold
from sklearn.linear_model import LinearRegression, Ridge, Lasso
from sklearn.ensemble import (RandomForestRegressor, GradientBoostingRegressor,
                              AdaBoostRegressor)
from sklearn.tree import DecisionTreeRegressor
from sklearn.svm import SVR
from sklearn.neighbors import KNeighborsRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.naive_bayes import GaussianNB
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.inspection import permutation_importance
from sklearn.metrics import (roc_curve, accuracy_score, f1_score, r2_score,
                             mean_squared_error, roc_auc_score, precision_score,
                             recall_score)
import shap

warnings.filterwarnings("ignore")

TRAIN_CSV = "train_en.csv"
VAL_CSV = "val_en.csv"
XLSX_OUT = "tennis_workflow_results_en.xlsx"
SEED = 42
N_SPLITS = 5
COLLINEAR_THRESH = 0.90
SHAP_BG = 50
SHAP_NSAMPLES = 150
PERMODEL_NSAMPLES = 100
TOP_K_CROSS = 10
np.random.seed(SEED)

def _init_plt():
    try:
        import logging
        logging.getLogger("matplotlib.font_manager").setLevel(logging.ERROR)
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
        for f in ["Microsoft YaHei", "SimHei", "PingFang SC", "Noto Sans CJK SC"]:
            if f in {ff.name for ff in font_manager.fontManager.ttflist}:
                plt.rcParams["font.sans-serif"] = [f]
                break
        plt.rcParams["axes.unicode_minus"] = False
        plt.rcParams["pdf.fonttype"] = 42
        plt.rcParams["ps.fonttype"] = 42
        return plt
    except Exception:
        return None

def _cjk_name():
    # English build: use matplotlib/seaborn default Latin font (no CJK)
    return None


def clean(df):
    df = df.copy()
    for c in df.columns:
        if df[c].dtype == "object":
            df[c] = df[c].astype(str).str.replace("、", "", regex=False)
        df[c] = pd.to_numeric(df[c], errors="coerce")
        if df[c].isna().all():
            df[c] = LabelEncoder().fit_transform(df[c].fillna("missing"))
        elif df[c].isna().any():
            df[c] = df[c].fillna(df[c].mean())
    return df

def youden_threshold(y_true, score):
    fpr, tpr, thr = roc_curve(y_true, score)
    return thr[np.argmax(tpr - fpr)]

def load_data():
    tr = clean(pd.read_csv(TRAIN_CSV))
    va = clean(pd.read_csv(VAL_CSV))
    y = tr.iloc[:, 0].astype(int).values
    X = tr.iloc[:, 1:]
    yv = va.iloc[:, 0].astype(int).values
    Xv = va.iloc[:, 1:]
    X = X.replace([np.inf, -np.inf], np.nan).fillna(X.mean())
    for c in Xv.columns:
        Xv[c] = Xv[c].replace([np.inf, -np.inf], np.nan).fillna(X[c].mean())
    return X.values.astype(float), y, Xv.values.astype(float), yv, list(X.columns)

def decorrelate(Xn, y, feat, thresh):
    n = len(feat)
    A = np.abs(np.corrcoef(Xn.T)) >= thresh
    seen = [False] * n
    clusters = []
    for i in range(n):
        if seen[i]:
            continue
        stack, comp = [i], []
        while stack:
            u = stack.pop()
            if seen[u]:
                continue
            seen[u] = True
            comp.append(u)
            for v in range(n):
                if v != u and A[u, v] and not seen[v]:
                    stack.append(v)
        clusters.append(comp)
    tcorr = np.array([abs(np.corrcoef(Xn[:, j], y)[0, 1]) for j in range(n)])
    keep, dropped = [], []
    for comp in clusters:
        if len(comp) > 1:
            rep = max(comp, key=lambda j: tcorr[j])
            keep.append(rep)
            for j in comp:
                if j != rep:
                    dropped.append({"Removed Feature": feat[j], "Representative kept": feat[rep],
                                    "Within-cluster r": round(abs(np.corrcoef(Xn[:, j], Xn[:, rep])[0, 1]), 3)})
        else:
            keep.append(comp[0])
    keep = sorted(set(keep))
    return keep, dropped

def _scaled(m):
    return Pipeline([("scaler", StandardScaler()), ("model", m)])

def make_models():
    return {
        "LR": _scaled(LinearRegression()),
        "Ridge": _scaled(Ridge(alpha=1.0)),
        "Lasso": _scaled(Lasso(alpha=0.1)),
        "RF": RandomForestRegressor(n_estimators=100, random_state=SEED),
        "GB": GradientBoostingRegressor(n_estimators=100, random_state=SEED),
        "SVR": _scaled(SVR(kernel="rbf", C=1.0, gamma="scale")),
        "AdaB": AdaBoostRegressor(n_estimators=100, random_state=SEED),
        "DT": DecisionTreeRegressor(random_state=SEED),
        "KNN": _scaled(KNeighborsRegressor(n_neighbors=5)),
        "MLP": _scaled(MLPRegressor(hidden_layer_sizes=(64,), max_iter=2000, random_state=SEED)),
        "GNB": _scaled(GaussianNB()),
        "LDA": _scaled(LinearDiscriminantAnalysis()),
    }

PROBA = {"GNB", "LDA"}
TREE = {"RF", "GB", "AdaB", "DT"}
LINEAR = {"LR", "Ridge", "Lasso"}

def mscore(nm, m, Xq):
    return m.predict_proba(Xq)[:, 1] if nm in PROBA else m.predict(Xq)

def compute_scores(X, y, Xv, names):
    skf = StratifiedKFold(N_SPLITS, shuffle=True, random_state=SEED)
    OOF = np.zeros((len(y), len(names)))
    for tri, vai in skf.split(X, y):
        mm = make_models()
        for j, nm in enumerate(names):
            OOF[vai, j] = mscore(nm, mm[nm].fit(X[tri], y[tri]), X[vai])
    mm = make_models()
    TEST = np.zeros((len(Xv), len(names)))
    for j, nm in enumerate(names):
        TEST[:, j] = mscore(nm, mm[nm].fit(X, y), Xv)
    return OOF, TEST

def composite_ablation(names, OOF, TEST, y, yv):
    n_models = len(names)
    total = (1 << n_models) - 1
    print(f"Enumerating all subsets ({total} total), selecting best per model count...")

    best_by_k = {}
    best_global_f1 = -1
    best_global_acc = -1
    best_global_combo = []

    count = 0
    for k in range(1, n_models + 1):
        best_f1_k, best_acc_k = -1, -1
        best_combo_k = None
        best_test_metrics = None

        for combo in combinations(names, k):
            count += 1
            if count % 500 == 0:
                print(f"  Processed {count}/{total} combos...")

            idx_list = [names.index(nm) for nm in combo]
            oof_avg = np.mean(OOF[:, idx_list], axis=1)
            thr = youden_threshold(y, oof_avg)
            oof_pred = (oof_avg >= thr).astype(int)
            f1_val = f1_score(y, oof_pred)
            acc_val = accuracy_score(y, oof_pred)

            if f1_val > best_f1_k or (f1_val == best_f1_k and acc_val > best_acc_k):
                best_f1_k = f1_val
                best_acc_k = acc_val
                best_combo_k = list(combo)

                test_avg = np.mean(TEST[:, idx_list], axis=1)
                test_pred = (test_avg >= thr).astype(int)
                test_acc = accuracy_score(yv, test_pred)
                test_f1 = f1_score(yv, test_pred)
                test_auc = roc_auc_score(yv, test_avg)
                best_test_metrics = {
                    'test_acc': round(test_acc, 4),
                    'test_f1': round(test_f1, 4),
                    'test_auc': round(test_auc, 4)
                }

        best_by_k[k] = {
            'combo': best_combo_k,
            'oof_f1': round(best_f1_k, 4),
            'oof_acc': round(best_acc_k, 4),
            **best_test_metrics
        }

        if best_f1_k > best_global_f1 or (best_f1_k == best_global_f1 and best_acc_k > best_global_acc):
            best_global_f1 = best_f1_k
            best_global_acc = best_acc_k
            best_global_combo = best_combo_k

    print(f"Enumeration done. Global best: {{{', '.join(best_global_combo)}}} (k={len(best_global_combo)})")
    print(f"  OOF_Acc={best_global_acc:.4f}, OOF_F1={best_global_f1:.4f}")

    rows = []
    for k in sorted(best_by_k.keys()):
        info = best_by_k[k]
        rows.append({
            "Num Models": k,
            "Model Combination": "+".join(info['combo']),
            "OOF_Acc": info['oof_acc'],
            "OOF_F1": info['oof_f1'],
            "Test_Acc": info['test_acc'],
            "Test_F1": info['test_f1'],
            "Test_AUC": info['test_auc'],
        })
    df = pd.DataFrame(rows)
    return df, best_global_combo

def compute_single_metrics(names, OOF, TEST, y, yv):
    rec = []
    for j, nm in enumerate(names):
        thr = youden_threshold(y, OOF[:, j])
        op = (OOF[:, j] >= thr).astype(int)
        tp = (TEST[:, j] >= thr).astype(int)
        try:
            tauc = roc_auc_score(yv, TEST[:, j])
        except Exception:
            tauc = np.nan
        rec.append({
            "Model": nm,
            "R2": round(r2_score(y, OOF[:, j]), 3),
            "MSE": round(mean_squared_error(y, OOF[:, j]), 3),
            "Accuracy": round(accuracy_score(y, op), 3),
            "F1": round(f1_score(y, op), 3),
            "Test_Accuracy": round(accuracy_score(yv, tp), 3),
            "Test_F1": round(f1_score(yv, tp), 3),
            "Test_AUC": round(tauc, 3)
        })
    return pd.DataFrame(rec)

def compute_classification_metrics(names, OOF, TEST, y, yv):
    rec = []
    for j, nm in enumerate(names):
        thr = youden_threshold(y, OOF[:, j])
        op = (OOF[:, j] >= thr).astype(int)
        tp = (TEST[:, j] >= thr).astype(int)
        rec.append({
            "Model": nm,
            "Accuracy": round(accuracy_score(yv, tp), 4),
            "Precision": round(precision_score(yv, tp, zero_division=0), 4),
            "Recall": round(recall_score(yv, tp, zero_division=0), 4),
            "F1": round(f1_score(yv, tp), 4)
        })
    return pd.DataFrame(rec)

def normalized_native_importance(X, y, Xv, yv, feat, names):
    """Uniform, model-agnostic feature importance for cross-model comparability.

    For every one of the 12 models we report PERMUTATION IMPORTANCE computed on
    the exact pipeline used throughout the study (non-tree models are wrapped in
    StandardScaler, so distance- and coefficient-based models are evaluated on
    standardized features rather than on raw, scale-confounded units). Importance
    is the mean decrease in score over n_repeats=30 random feature permutations on
    the training set; negative values (permuting the feature does not hurt) are
    set to 0, and each model's vector is normalized to sum to 1. Using a single
    method for all models makes the rows directly comparable and avoids mixing
    impurity gain, |coef| and permutation importance, which are not commensurable.
    """
    N_REPEATS = 30
    mm = make_models()
    results = {}
    for nm in names:
        m = mm[nm].fit(X, y)
        imp = permutation_importance(m, X, y, n_repeats=N_REPEATS,
                                     random_state=SEED).importances_mean.copy()
        imp = np.clip(imp, 0.0, None)                       # negative permutation importance -> 0
        imp_norm = imp / imp.sum() if imp.sum() > 0 else np.ones(len(imp)) / len(imp)
        results[nm] = dict(zip(feat, [round(v, 6) for v in imp_norm]))
    return pd.DataFrame(results).T

def feature_ablation(names_sel, X, y, Xv, yv, feat):
    from sklearn.metrics import accuracy_score, f1_score, roc_auc_score

    skf = StratifiedKFold(N_SPLITS, shuffle=True, random_state=SEED)

    def oof_scores(Xfull):
        # 5-fold out-of-fold ensemble scores on the TRAINING set (n=199).
        # Folds depend only on y + random_state, so they are identical for any column subset.
        oof = np.zeros(len(y))
        for tri, vai in skf.split(Xfull, y):
            mmt = make_models()
            for nm in names_sel:
                mmt[nm].fit(Xfull[tri], y[tri])
            oof[vai] = np.mean([mscore(nm, mmt[nm], Xfull[vai]) for nm in names_sel], axis=0)
        return oof

    # ---- baselines ----
    oof = oof_scores(X)
    thr = youden_threshold(y, oof)
    base_oof_auc = roc_auc_score(y, oof)

    mm = make_models()
    fitted = {nm: mm[nm].fit(X, y) for nm in names_sel}

    def ensemble_predict(Xq):
        return np.mean([mscore(nm, fitted[nm], Xq) for nm in names_sel], axis=0)

    y_score_base = ensemble_predict(Xv)
    base_pred = (y_score_base >= thr).astype(int)
    base_acc = accuracy_score(yv, base_pred)
    base_f1 = f1_score(yv, base_pred)
    base_test_auc = roc_auc_score(yv, y_score_base)

    results = []
    for j in range(len(feat)):
        mask = [c for c in range(len(feat)) if c != j]

        # test-set ablation
        mm_abl = make_models()
        fitted_abl = {nm: mm_abl[nm].fit(X[:, mask], y) for nm in names_sel}

        def pred_abl(Xq):
            return np.mean([mscore(nm, fitted_abl[nm], Xq) for nm in names_sel], axis=0)

        y_score_abl = pred_abl(Xv[:, mask])
        pred_abl_lbl = (y_score_abl >= thr).astype(int)
        acc_abl = accuracy_score(yv, pred_abl_lbl)
        f1_abl = f1_score(yv, pred_abl_lbl)
        test_auc_abl = roc_auc_score(yv, y_score_abl)

        # OOF ablation (n=199, threshold-free)
        oof_auc_abl = roc_auc_score(y, oof_scores(X[:, mask]))

        results.append({
            "Feature": feat[j],
            "AUC_Drop(OOF)": round(base_oof_auc - oof_auc_abl, 4),
            "AUC_Drop(test)": round(base_test_auc - test_auc_abl, 4),
            "Acc_Drop": round(base_acc - acc_abl, 4),
            "F1_Drop": round(base_f1 - f1_abl, 4),
            "Baseline AUC(OOF)": round(base_oof_auc, 4),
            "Ablated AUC(OOF)": round(oof_auc_abl, 4),
            "Baseline AUC(test)": round(base_test_auc, 4),
            "Ablated AUC(test)": round(test_auc_abl, 4),
            "Baseline Acc": round(base_acc, 4),
            "Ablated Acc": round(acc_abl, 4),
            "Baseline F1": round(base_f1, 4),
            "Ablated F1": round(f1_abl, 4)
        })

    return pd.DataFrame(results).sort_values("AUC_Drop(OOF)", ascending=False)

# ===================== SHAPCorr =====================
def fit_ensemble(names_sel, X, y):
    mm = make_models()
    fitted = {nm: mm[nm].fit(X, y) for nm in names_sel}
    def predict(Xq):
        return np.mean([mscore(nm, fitted[nm], Xq) for nm in names_sel], axis=0)
    return fitted, predict

def ensemble_shap(predict, X, Xv, feat, top_k=6):
    bg = shap.sample(X, min(SHAP_BG, len(X)), random_state=SEED)
    expl = shap.KernelExplainer(predict, bg)
    sv = expl.shap_values(Xv, nsamples=SHAP_NSAMPLES, silent=True)
    base = float(np.array(expl.expected_value).ravel()[0])
    msh = np.abs(sv).mean(axis=0)
    order = np.argsort(-msh)

    n_select = max(top_k, int(len(feat) * 0.2))
    n_select = min(n_select, len(feat))
    top_indices = order[:n_select]

    df = pd.DataFrame({
        "Feature": [feat[i] for i in order],
        "mean_abs_SHAP": [round(msh[i], 4) for i in order],
        "Direction (vs outcome)": ["Positive" if np.corrcoef(Xv[:, i], sv[:, i])[0, 1] > 0 else "Negative" for i in order]
    })
    df.insert(0, "Rank", range(1, len(order) + 1))

    top_feat_names = [feat[i] for i in top_indices]
    return sv, base, df, top_feat_names, top_indices

def plot_roc_curves(names, X, y, Xv, yv, fname="figure1_12models_ROC.pdf"):
    plt = _init_plt()
    if plt is None:
        return

    mm = make_models()
    colors = plt.cm.tab10(np.linspace(0, 1, len(names)))

    plt.figure(figsize=(9, 7))
    for j, nm in enumerate(names):
        m = mm[nm].fit(X, y)
        y_score = mscore(nm, m, Xv)
        fpr, tpr, _ = roc_curve(yv, y_score)
        auc_val = roc_auc_score(yv, y_score)
        plt.plot(fpr, tpr, color=colors[j], lw=1.5, label=f"{nm} (AUC={auc_val:.3f})")

    plt.plot([0, 1], [0, 1], "k--", lw=0.8, alpha=0.6)
    plt.xlim([0, 1])
    plt.ylim([0, 1.05])
    plt.xlabel("False Positive Rate (FPR)")
    plt.ylabel("True Positive Rate (TPR)")
    plt.title("ROC Curves of 12 Single Models")
    plt.legend(loc="lower right", fontsize=7, ncol=2)
    plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(fname, format="pdf")
    plt.close()
    print("  Saved:", fname)

def plot_shap_beeswarm(sv, Xv, feat, top_indices, fname="figure2_SHAP_beeswarm.pdf"):
    plt = _init_plt()
    if plt is None:
        return

    sv_sub = sv[:, top_indices]
    Xv_sub = Xv[:, top_indices]
    feat_sub = [feat[i] for i in top_indices]

    plt.figure(figsize=(10, 7))
    shap.summary_plot(sv_sub, Xv_sub, feature_names=feat_sub, plot_type="dot",
                      max_display=len(feat_sub), show=False)
    for _ax in plt.gcf().axes:
        for _c in _ax.collections:
            _c.set_rasterized(False)
    plt.title(f"SHAP Summary (Top {len(feat_sub)} indicators)")
    plt.tight_layout()
    plt.savefig(fname, format="pdf", bbox_inches="tight")
    plt.close()
    print("  Saved:", fname)

def plot_shap_bar(sv, Xv, feat, top_indices, fname="figure3_SHAP_bar.pdf"):
    plt = _init_plt()
    if plt is None:
        return

    sv_sub = sv[:, top_indices]
    Xv_sub = Xv[:, top_indices]
    feat_sub = [feat[i] for i in top_indices]

    plt.figure(figsize=(9, 6))
    shap.summary_plot(sv_sub, Xv_sub, feature_names=feat_sub, plot_type="bar",
                      max_display=len(feat_sub), show=False)
    plt.title(f"SHAP Importance (Top {len(feat_sub)} indicators)")
    plt.tight_layout()
    plt.savefig(fname, format="pdf", bbox_inches="tight")
    plt.close()
    print("  Saved:", fname)

def plot_shap_waterfall(sv, base, Xv, yv, feat, top_indices, ens_score, fname_prefix="figure4_SHAP_waterfall"):
    plt = _init_plt()
    if plt is None:
        return

    win_idx = np.where(yv == 1)[0]
    loss_idx = np.where(yv == 0)[0]

    samples = []
    if len(win_idx) > 0:
        si = int(win_idx[np.argsort(ens_score[win_idx])[len(win_idx) // 2]])
        samples.append(("Typical Win", si))
    if len(loss_idx) > 0:
        si = int(loss_idx[np.argsort(ens_score[loss_idx])[len(loss_idx) // 2]])
        samples.append(("Typical Loss", si))

    for label, si in samples:
        plt.figure(figsize=(10, 6))
        shap.waterfall_plot(
            shap.Explanation(
                values=sv[si],
                base_values=base,
                data=Xv[si],
                feature_names=feat
            ), max_display=len(top_indices) + 1, show=False
        )
        plt.title(f"{label}: actual={yv[si]}, pred={ens_score[si]:.3f}, base={base:.3f}")
        plt.tight_layout()
        tag = "win" if "Win" in label else "loss"
        fname = f"{fname_prefix}_{tag}.pdf"
        plt.savefig(fname, format="pdf", bbox_inches="tight")
        plt.close()
        print("  Saved:", fname)

def plot_shap_rank_distribution(per_rank, model_names, top_feat_names, n_feat, fname, title):
    import seaborn as sns
    cjk = _cjk_name()
    sns.set_theme(style="white", **({"font": cjk} if cjk else {}))
    plt = _init_plt()
    if plt is None:
        return
    if cjk:
        plt.rcParams["font.family"] = cjk
    plt.rcParams.update({"pdf.fonttype": 42, "ps.fonttype": 42})
    rows = [{"Feature": f, "Model": nm, "Rank": per_rank[nm][f]}
            for f in top_feat_names for nm in model_names]
    dfl = pd.DataFrame(rows)
    order = dfl.groupby("Feature")["Rank"].mean().sort_values().index.tolist()
    fig, ax = plt.subplots(figsize=(7.6, 4.8), facecolor="white")
    sns.boxplot(data=dfl, x="Feature", y="Rank", order=order, ax=ax, width=0.56,
                showcaps=True, fliersize=0,
                boxprops=dict(facecolor="#D9E3F0", edgecolor="#5B728A", linewidth=1.0),
                whiskerprops=dict(color="#5B728A", linewidth=1.0),
                capprops=dict(color="#5B728A", linewidth=1.0),
                medianprops=dict(color="#1F2D3A", linewidth=1.4))
    sns.stripplot(data=dfl, x="Feature", y="Rank", order=order, ax=ax, color="#243746",
                  size=5, jitter=0.18, alpha=0.75, zorder=3)
    ax.set_ylim(n_feat + 1, 0)
    ax.set_title(title, pad=10)
    ax.set_xlabel(""); ax.set_ylabel("SHAP rank (1 = most important)")
    for lb in ax.get_xticklabels():
        lb.set_rotation(30); lb.set_ha("right")
    ax.yaxis.grid(True, linestyle=(0, (2, 3)), linewidth=0.6, color="#D9DEE5"); ax.xaxis.grid(False)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.tight_layout(); fig.savefig(fname, bbox_inches="tight", facecolor="white"); plt.close(fig)
    print("  Saved:", fname)

def plot_shap_cross_model_coverage(per_rank, model_names, top_feat_names, top_k, fname, title):
    import seaborn as sns
    cjk = _cjk_name()
    sns.set_theme(style="white", **({"font": cjk} if cjk else {}))
    freq = {f: sum(1 for nm in model_names if per_rank[nm].get(f, 999) <= top_k) for f in top_feat_names}
    freq_series = pd.Series(freq).sort_values(ascending=True)
    plt = _init_plt()
    if plt is None:
        return None
    if cjk:
        plt.rcParams["font.family"] = cjk
    plt.rcParams.update({"font.size": 10, "axes.titlesize": 12, "axes.labelsize": 10.5,
                         "pdf.fonttype": 42, "ps.fonttype": 42})
    fig, ax = plt.subplots(figsize=(6.6, 4.8), facecolor="white")
    y_pos = np.arange(len(freq_series)); x_vals = freq_series.values
    for yy, xx in zip(y_pos, x_vals):
        ax.hlines(y=yy, xmin=0, xmax=xx, color="#CCD6E0", linewidth=1.2, zorder=1)
    sizes = 40 + (freq_series.max() - freq_series) * 18
    ax.scatter(x_vals, y_pos, s=sizes.values, facecolor="#6E8FA8", edgecolor="#2F4858",
               linewidth=0.9, alpha=0.95, zorder=3)
    ax.set_yticks(y_pos)
    ax.set_yticklabels([str(f)[:33] + ("..." if len(str(f)) > 34 else "") for f in freq_series.index])
    ax.set_xlabel(f"Models with feature in SHAP top-{top_k} (of {len(model_names)})")
    ax.set_ylabel(""); ax.set_title(title, pad=10, weight="semibold")
    ax.set_xlim(0, max(x_vals) + 0.8); ax.set_xticks(range(0, int(max(x_vals)) + 1))
    ax.xaxis.grid(True, linestyle=(0, (2, 3)), linewidth=0.6, color="#D9DEE5"); ax.yaxis.grid(False)
    for xx, yy in zip(x_vals, y_pos):
        ax.text(xx + 0.08, yy, f"{int(xx)}", va="center", ha="left", fontsize=9, color="#2F4858")
    for sp in ["top", "right", "left"]:
        ax.spines[sp].set_visible(False)
    plt.tight_layout(); plt.savefig(fname, bbox_inches="tight", facecolor="white"); plt.close()
    print("  Saved:", fname)
    return pd.DataFrame({"Top Indicator": list(freq_series.index),
                         f"Models in top-{top_k}": list(freq_series.values),
                         "Model Group": f"{len(model_names)} models"})

def main():
    X0, y, Xv0, yv, feat0 = load_data()
    print(f"Raw data: train {X0.shape}, test {Xv0.shape}, {len(feat0)} features")

    normality_results = []
    for i, fn in enumerate(feat0):
        stat, p = shapiro(X0[:, i])
        normality_results.append({
            "Feature": fn,
            "Shapiro-Wilk W": round(stat, 4),
            "p-value": round(p, 4),
            "Normality(p<0.05)": "No (non-normal)" if p < 0.05 else "Yes (normal)",
            "Corr Method": "Spearman" if p < 0.05 else "Pearson"
        })
    df_normality = pd.DataFrame(normality_results)

    corr_results = []
    for i, fn in enumerate(feat0):
        is_normal = normality_results[i]["Normality(p<0.05)"] == "Yes (normal)"
        if is_normal:
            r, p_val = pearsonr(X0[:, i], y)
            method = "Pearson"
        else:
            r, p_val = spearmanr(X0[:, i], y)
            method = "Spearman"
        corr_results.append({
            "Feature": fn,
            "Corr Method": method,
            "Corr Coef": round(r, 4),
            "p-value": round(p_val, 4),
            "Significance (p<0.05)": "Significant" if p_val < 0.05 else "n.s.",
            "Direction": "Positive" if r > 0 else "Negative"
        })
    df_corr = pd.DataFrame(corr_results).sort_values("Corr Coef", key=abs, ascending=False)

    keep, dropped = decorrelate(X0, y, feat0, COLLINEAR_THRESH)
    X, Xv, feat = X0[:, keep], Xv0[:, keep], [feat0[i] for i in keep]
    print(f"De-collinearization (|r|>={COLLINEAR_THRESH}): removed {len(dropped)}, kept {len(feat)}")

    names = list(make_models().keys())
    OOF, TEST = compute_scores(X, y, Xv, names)

    df_single = compute_single_metrics(names, OOF, TEST, y, yv)

    df_class_metrics = compute_classification_metrics(names, OOF, TEST, y, yv)

    df_ablation, best_combo = composite_ablation(names, OOF, TEST, y, yv)

    sel = best_combo
    print(f"Final ensemble: {sel}")

    skf = StratifiedKFold(N_SPLITS, shuffle=True, random_state=SEED)
    oof_sel = np.zeros(len(y))
    for tri, vai in skf.split(X, y):
        mm = make_models()
        for nm in sel:
            mm[nm].fit(X[tri], y[tri])
        oof_sel[vai] = np.mean([mscore(nm, mm[nm], X[vai]) for nm in sel], axis=0)
    thr_sel = youden_threshold(y, oof_sel)

    fitted_sel = make_models()
    for nm in sel:
        fitted_sel[nm].fit(X, y)
    def predict_sel(Xq):
        return np.mean([mscore(nm, fitted_sel[nm], Xq) for nm in sel], axis=0)
    ens_te = predict_sel(Xv)

    df_native_imp = normalized_native_importance(X, y, Xv, yv, feat, names)

    sv, base, df_shap, top_feat_names, top_indices = ensemble_shap(predict_sel, X, Xv, feat, top_k=max(6, int(len(feat) * 0.2)))
    print(f"Top features (n={len(top_indices)}): {top_feat_names}")

    bg_shap = shap.sample(X, min(SHAP_BG, len(X)), random_state=SEED)
    mm_shap = make_models()
    shap_rank_dict = {}
    for nm in names:
        m = mm_shap[nm].fit(X, y)
        try:
            if nm in TREE:
                sv_m = shap.TreeExplainer(m).shap_values(Xv, check_additivity=False)
            elif nm in LINEAR:
                sc = m.named_steps["scaler"]
                est = m.named_steps["model"]
                sv_m = shap.LinearExplainer(est, sc.transform(bg_shap)).shap_values(sc.transform(Xv))
            else:
                f = (lambda A, _m=m: _m.predict_proba(A)[:, 1]) if nm in PROBA else m.predict
                sv_m = shap.KernelExplainer(f, bg_shap).shap_values(Xv, nsamples=PERMODEL_NSAMPLES, silent=True)
            sv_m = np.array(sv_m)
            if sv_m.ndim == 3:
                sv_m = sv_m[..., -1]
            msh_m = np.abs(sv_m).mean(axis=0)
        except Exception:
            msh_m = np.zeros(len(feat))
        order = np.argsort(-msh_m)
        shap_rank_dict[nm] = {feat[j]: int(np.where(order == j)[0][0]) + 1 for j in range(len(feat))}

    shap_rank_df = pd.DataFrame({nm: {f: shap_rank_dict[nm].get(f, "-") for f in top_feat_names} for nm in names}).T
    shap_rank_df.index.name = "Model"

    plot_roc_curves(names, X, y, Xv, yv)
    plot_shap_beeswarm(sv, Xv, feat, top_indices)
    plot_shap_bar(sv, Xv, feat, top_indices)
    plot_shap_waterfall(sv, base, Xv, yv, feat, top_indices, ens_te)
    plot_shap_rank_distribution(shap_rank_dict, names, top_feat_names, len(feat),
                                "figure5_SHAP_distribution_12models.pdf", "SHAP Rank Distribution (12 single models)")
    plot_shap_rank_distribution(shap_rank_dict, sel, top_feat_names, len(feat),
                                "figure5_SHAP_distribution_selected.pdf", f"SHAP Rank Distribution ({len(sel)} selected members)")
    df_cov12 = plot_shap_cross_model_coverage(shap_rank_dict, names, top_feat_names, TOP_K_CROSS,
                                "figure6_SHAP_cross_model_coverage_12models.pdf", "Cross-Model Coverage (12 single models)")
    df_covsel = plot_shap_cross_model_coverage(shap_rank_dict, sel, top_feat_names, TOP_K_CROSS,
                                "figure6_SHAP_cross_model_coverage_selected.pdf", f"Cross-Model Coverage ({len(sel)} selected members)")

    # perturb the (global-RNG-dependent) SHAP sheets. Ablation does not feed SHAP anyway.
    df_feat_abl = feature_ablation(sel, X, y, Xv, yv, feat)

    print("\n===== Writing Excel =====")
    df_normality = df_normality.rename(columns={"Normality(p<0.05)": "Normal?"})
    df_cov = pd.concat([df_cov12, df_covsel], ignore_index=True) if df_cov12 is not None else pd.DataFrame()
    with pd.ExcelWriter(XLSX_OUT, engine="openpyxl") as w:
        df_normality.to_excel(w, sheet_name="1-Normality", index=False)
        df_corr.to_excel(w, sheet_name="2-Correlation", index=False)
        df_single.to_excel(w, sheet_name="3-Single Model Metrics", index=False)
        df_class_metrics.to_excel(w, sheet_name="4-Classification Metrics", index=False)
        df_ablation.to_excel(w, sheet_name="5-Ensemble Ablation", index=False)
        df_native_imp.to_excel(w, sheet_name="6-Native Importance (norm)")
        df_feat_abl.to_excel(w, sheet_name="7-Feature Ablation", index=False)
        df_shap.to_excel(w, sheet_name="8-Ensemble SHAP (All)", index=False)
        shap_rank_df.to_excel(w, sheet_name="9-Single-Model SHAP Rank")
        df_cov.to_excel(w, sheet_name="10-Cross-Model Coverage", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = load_workbook(XLSX_OUT); fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        for c in ws[1]:
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = fill; c.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            wd = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(wd + 2, 10), 46)
        ws.freeze_panes = "A2"
    wb.save(XLSX_OUT)

    print(f"  Saved: {XLSX_OUT}")
    print("\n[Done]")
    print(f"Final ensemble: {sel}")
    print("Output files:")
    print(f"  {XLSX_OUT}")
    print("  figure1_12models_ROC.pdf")
    print("  figure2_SHAP_beeswarm.pdf  figure3_SHAP_bar.pdf")
    print("  figure4_SHAP_waterfall_win/loss.pdf")
    print("  figure5_SHAP_distribution_12models.pdf  figure5_SHAP_distribution_selected.pdf")
    print("  figure6_SHAP_cross_model_coverage_12models.pdf  figure6_SHAP_cross_model_coverage_selected.pdf")

if __name__ == "__main__":
    main()